import plotly.graph_objects as go
import pandas as pd
import streamlit as st
from config.rutas import EXCEL_PATH
import time
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

def cargar_excel():
    try:
        prog_df = pd.read_excel(EXCEL_PATH)
        st.success("Datos cargados exitosamente")
    except Exception as e:
        st.error(f"Error al cargar los datos: {str(e)}")
    return prog_df

def agregar_datos(df, datos):
    pass

def exportar_excel(df,file_name):
    now = time.strftime("%Y%m%d-%H%M%S")
    df.to_excel(file_name, index=False)
    wb = load_workbook(file_name)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column
    col_final = get_column_letter(max_col)

    tabla = Table(displayName=f"Tabla{now}", ref=f"A1:{col_final}{max_row}")
    estilo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)

    wb.save(file_name)
    st.markdown(f"✅ Archivo creado: {file_name}")
    
def exportar_pdf(df, archivo_pdf):
    import plotly.io as pio

    fig = go.Figure(data=[go.Table(
        header=dict(
            values=list(df.columns),
            fill_color='lightgray',
            align='left',
            font=dict(size=10, color='black')
        ),
        cells=dict(
            values=[df[col] for col in df.columns],
            fill_color='white',
            align='left',
            font=dict(size=9, color='black'),
            height=20
        )
    )])

    fig.update_layout(
        margin=dict(l=20, r=20, t=40, b=20),
        width=2400,  # más ancho para columnas más anchas
        height=3000  # alto depende del número de filas
    )

    # Exportar a PDF
    pio.write_image(fig, archivo_pdf, format='pdf', width=2400, height=3000)
    print(f"✅ PDF generado en {archivo_pdf}")

